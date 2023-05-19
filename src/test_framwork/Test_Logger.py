import os
import csv

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from scapy.utils import PcapWriter

class TestLogger:
    def __init__(self, file_name=None):
        self._file_name = file_name
        self.logf = None
        self._field_names = []

    def create_csv(self, field_names):
        self._field_names = field_names
        data_dir = './result'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        with open(self._file_name, 'a+') as out_file:
            writer = csv.DictWriter(out_file, fieldnames = field_names)
            writer.writeheader()

    def write2csv(self, info_dicts):
        with open(self._file_name, 'a+') as out_file:
            writer = csv.DictWriter(out_file, fieldnames = self._field_names)
            writer.writerow(info_dicts)

    def creat_binf_sct2(self, file_name, sn_num, test_time):
        data_dir = f'./data/Packet_ODR_test_data/{sn_num}_{test_time}'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        self.logf = open(file_name, 'wb')

    def cerat_binf_sct5(self, file_name):
        data_dir = f'./data/Packet_long_term_test_data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        self.logf = open(file_name, 'wb')

    def creat_binf_sct7(self, file_name):
        data_dir = f'./data/static_test_data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        self.logf = open(file_name, 'wb')

    def write2bin(self, data):
        if data is not None:
            self.logf.write(data)

    def write2pcap(self, data, logf):
        data_dir = f'./data/Packet_long_term_test_data'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        pktdump = PcapWriter(logf, append=True, sniff=True)
        pktdump.write(data)

    def backup_log(self, log_str=None):
        data_dir = f'./log'
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
        if log_str is not None:
            logf = open(f'./log/backup_log.txt', 'a+')
            logf.write(log_str + '\n')
            logf.close()

class xlsx_factory: 
    def __init__(self, file_name, mode='static') -> None:
        if mode == 'static':
            self.wb = load_workbook('./case/INS401_ETHERNET_INTERFACE_FUNCTION_TEST_CASE_v3.0.xlsx')
        elif mode == 'dynamic':
            self.wb = load_workbook(file_name)
        self.ws = self.wb.active
        self.ts0 = self.wb.worksheets[0]
        self.ts1 = self.wb.worksheets[1]
        self.row_dict = {}

    def sheet_ready(self):
        i = 0
        for row in self.ts1.rows:
            i += 1
            row_value = []
            for idx in row:
                value = idx.value
                row_value.append(value)
            self.row_dict.update({i: row_value})
        # print(self.row_dict)

        for i in range(len(self.row_dict)):
            title = self.row_dict[i + 1][0]
            if title == 'E1':
                E1_case_start_line = i + 2
            if title == 'E2':
                E2_case_start_line = i + 2
            if title == 'E3':
                E3_case_start_line = i + 2
            if title == 'E4':
                E4_case_start_line = i + 2
            if title == 'E5':
                E5_case_start_line = i + 2
            if title == 'E6':
                E6_case_start_line = i + 2
            if title == 'E7':
                E7_case_start_line = i + 2
            if title == 'E8':
                E8_case_start_line = i + 2
            if title == 'E9':
                E9_case_start_line = i + 2
            if title == 'E10':
                E10_case_start_line = i + 2
            if title == 'E11':
                E11_case_start_line = i + 2
            if title == 'E12':
                E12_case_start_line = i + 2
            if title == 'E13':
                E13_case_start_line = i + 2
            if title == 'E14':
                E14_case_start_line = i + 2
            if title == 'E15':
                E15_case_start_line = i + 2
            if title == 'E16':
                E16_case_start_line = i + 2
            if title == 'E17':
                E17_case_start_line = i + 2
            if title == 'E18':
                E18_case_start_line = i + 2

        test_case_range = {
            'E1': [E1_case_start_line, E2_case_start_line-2],
            'E2': [E2_case_start_line, E3_case_start_line-2],
            'E3': [E3_case_start_line, E4_case_start_line-2],
            'E4': [E4_case_start_line, E5_case_start_line-2],
            'E5': [E5_case_start_line, E6_case_start_line-2],
            'E6': [E6_case_start_line, E7_case_start_line-2],
            'E7': [E7_case_start_line, E8_case_start_line-2],
            'E8': [E8_case_start_line, E9_case_start_line-2],
            'E9': [E9_case_start_line, E10_case_start_line-2],
            'E10': [E10_case_start_line, E11_case_start_line-2],
            'E11': [E11_case_start_line, E12_case_start_line-2],
            'E12': [E12_case_start_line, E13_case_start_line-2],
            'E13': [E13_case_start_line, E14_case_start_line-2],
            'E14': [E14_case_start_line, E15_case_start_line-2],
            'E15': [E15_case_start_line, E16_case_start_line-2],
            'E16': [E16_case_start_line, E17_case_start_line-2],
            'E17': [E17_case_start_line, E18_case_start_line-2],
            'E18': [E18_case_start_line, len(self.row_dict)]
        }
    
        test_actual_val_field = {
            'E1': self.ts1[f'E{test_case_range["E1"][0]}:E{test_case_range["E1"][1]}'],
            'E2': self.ts1[f'E{test_case_range["E2"][0]}:E{test_case_range["E2"][1]}'],
            'E3': self.ts1[f'E{test_case_range["E3"][0]}:E{test_case_range["E3"][1]}'],
            'E4': self.ts1[f'E{test_case_range["E4"][0]}:E{test_case_range["E4"][1]}'],
            'E5': self.ts1[f'E{test_case_range["E5"][0]}:E{test_case_range["E5"][1]}'],
            'E6': self.ts1[f'E{test_case_range["E6"][0]}:E{test_case_range["E6"][1]}'],
            'E7': self.ts1[f'E{test_case_range["E7"][0]}:E{test_case_range["E7"][1]}'],
            'E8': self.ts1[f'E{test_case_range["E8"][0]}:E{test_case_range["E8"][1]}'],
            'E9': self.ts1[f'E{test_case_range["E9"][0]}:E{test_case_range["E9"][1]}'],
            'E10': self.ts1[f'E{test_case_range["E10"][0]}:E{test_case_range["E10"][1]}'],
            'E11': self.ts1[f'E{test_case_range["E11"][0]}:E{test_case_range["E11"][1]}'],
            'E12': self.ts1[f'E{test_case_range["E12"][0]}:E{test_case_range["E12"][1]}'],
            'E13': self.ts1[f'E{test_case_range["E13"][0]}:E{test_case_range["E13"][1]}'],
            'E14': self.ts1[f'E{test_case_range["E14"][0]}:E{test_case_range["E14"][1]}'],
            'E15': self.ts1[f'E{test_case_range["E15"][0]}:E{test_case_range["E15"][1]}'],
            'E16': self.ts1[f'E{test_case_range["E16"][0]}:E{test_case_range["E16"][1]}'],
            'E17': self.ts1[f'E{test_case_range["E17"][0]}:E{test_case_range["E17"][1]}'],
            'E18': self.ts1[f'E{test_case_range["E18"][0]}:E{test_case_range["E18"][1]}']
        }

        test_result_field = {
            'E1': self.ts1[f'F{test_case_range["E1"][0]}:F{test_case_range["E1"][1]}'],
            'E2': self.ts1[f'F{test_case_range["E2"][0]}:F{test_case_range["E2"][1]}'],
            'E3': self.ts1[f'F{test_case_range["E3"][0]}:F{test_case_range["E3"][1]}'],
            'E4': self.ts1[f'F{test_case_range["E4"][0]}:F{test_case_range["E4"][1]}'],
            'E5': self.ts1[f'F{test_case_range["E5"][0]}:F{test_case_range["E5"][1]}'],
            'E6': self.ts1[f'F{test_case_range["E6"][0]}:F{test_case_range["E6"][1]}'],
            'E7': self.ts1[f'F{test_case_range["E7"][0]}:F{test_case_range["E7"][1]}'],
            'E8': self.ts1[f'F{test_case_range["E8"][0]}:F{test_case_range["E8"][1]}'],
            'E9': self.ts1[f'F{test_case_range["E9"][0]}:F{test_case_range["E9"][1]}'],
            'E10': self.ts1[f'F{test_case_range["E10"][0]}:F{test_case_range["E10"][1]}'],
            'E11': self.ts1[f'F{test_case_range["E11"][0]}:F{test_case_range["E11"][1]}'],
            'E12': self.ts1[f'F{test_case_range["E12"][0]}:F{test_case_range["E12"][1]}'],
            'E13': self.ts1[f'F{test_case_range["E13"][0]}:F{test_case_range["E13"][1]}'],
            'E14': self.ts1[f'F{test_case_range["E14"][0]}:F{test_case_range["E14"][1]}'],
            'E15': self.ts1[f'F{test_case_range["E15"][0]}:F{test_case_range["E15"][1]}'],
            'E16': self.ts1[f'F{test_case_range["E16"][0]}:F{test_case_range["E16"][1]}'],
            'E17': self.ts1[f'F{test_case_range["E17"][0]}:F{test_case_range["E17"][1]}'],
            'E18': self.ts1[f'F{test_case_range["E18"][0]}:F{test_case_range["E18"][1]}']
        }

        return test_actual_val_field, test_result_field

    # def edtior(self):

    def change_val(self, field, caseID, val):
        field[caseID][0].value = val

    def savef(self, file_name):
        self.wb.save(file_name)

    def set_cell_color(self, field, caseID):
        fille = PatternFill('solid', fgColor='c00000')
        cell = field[caseID][0]
        cell.fill = fille

    
